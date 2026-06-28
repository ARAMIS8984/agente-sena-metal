import os, json, logging, requests, io
from flask import Flask, request, jsonify
from datetime import datetime
from generar_pdf import generar_pdf_llamado_oficial, FALTAS_DISC, FALTAS_ACAD, PROGRAMAS_NOMBRES
from generar_plan_concertado import generar_pdf_plan_concertado, generar_zip_plan_concertado
from catalogo import CATALOGO

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
app = Flask(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN")
GROQ_API_KEY   = os.environ.get("GROQ_API_KEY")
TELEGRAM_API   = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
GROQ_URL       = "https://api.groq.com/openai/v1/chat/completions"

estados = {}
conversaciones = {}

SYSTEM_ANALISIS = """Eres el Asistente SENA Metalmecánica del Centro Nacional Colombo Alemán, Barranquilla. Eres como un colega experimentado que ayuda a los instructores.

Cuando recibes una novedad, analizas la situación de forma natural y humana. Respondes SIEMPRE con este formato:

---
🎓 *ANÁLISIS DE NOVEDAD*
*Aprendiz:* [nombre]
*Ficha:* [codigo o no especificada]

📋 *SITUACIÓN:*
[Análisis breve y humano de la situación]

⚖️ *NORMA APLICABLE:*
[Artículo del Acuerdo 0009 de 2024 con mayor peso argumentativo]

✅ *PROCEDIMIENTO RECOMENDADO:*
1. [Paso 1]
2. [Paso 2]
3. [Paso 3]

💡 *MI SUGERENCIA:*
[Tipo de documento recomendado]
---

REGLAS DEL ACUERDO 0009 DE 2024:
- Art. 8 lit. d: Asistir puntualmente — 3+ ausencias → Llamado tipo 1
- Art. 16: 5+ ausencias → Llamado tipo 2 + Plan de Mejoramiento
- Art. 22: Nota mínima 3.0 — menor → Plan de Mejoramiento
- Art. 28: 60% mínimo de evidencias
- Art. 20: Consumo sustancias → causal cancelación inmediata

Responde en español, profesional y directo. Máximo 200 palabras."""

SYSTEM_REGLAMENTO = """Eres experto en el Acuerdo 0009 de 2024 del Reglamento del Aprendiz SENA.
Respondes preguntas citando el artículo exacto y explicando su aplicación práctica.
Responde en español, claro y conciso."""

def get_estado(chat_id):
    if chat_id not in estados:
        estados[chat_id] = {"modo": "menu"}
    return estados[chat_id]

def set_estado(chat_id, datos):
    estados[chat_id] = datos

def send_message(chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if reply_markup:
        payload["reply_markup"] = json.dumps(reply_markup)
    try:
        r = requests.post(f"{TELEGRAM_API}/sendMessage", json=payload, timeout=30)
        if r.status_code != 200:
            payload2 = {"chat_id": chat_id, "text": text.replace("*","").replace("_","").replace("`","")}
            if reply_markup:
                payload2["reply_markup"] = json.dumps(reply_markup)
            requests.post(f"{TELEGRAM_API}/sendMessage", json=payload2, timeout=30)
    except Exception as e:
        logger.error(f"Error send: {e}")

def send_document(chat_id, doc_bytes, filename, caption=""):
    try:
        r = requests.post(
            f"{TELEGRAM_API}/sendDocument",
            data={"chat_id": chat_id, "caption": caption},
            files={"document": (filename, doc_bytes, "application/pdf" if filename.endswith('.pdf') else "application/zip")},
            timeout=60
        )
        logger.info(f"Send doc: {r.status_code}")
    except Exception as e:
        logger.error(f"Error send doc: {e}")

def send_typing(chat_id):
    try:
        requests.post(f"{TELEGRAM_API}/sendChatAction",
                      json={"chat_id": chat_id, "action": "typing"}, timeout=10)
    except: pass

def answer_callback(callback_id):
    try:
        requests.post(f"{TELEGRAM_API}/answerCallbackQuery",
                      json={"callback_query_id": callback_id}, timeout=10)
    except: pass

def menu_principal(chat_id):
    set_estado(chat_id, {"modo": "menu"})
    conversaciones.pop(chat_id, None)
    send_message(chat_id,
        "👋 Bienvenido al *Agente SENA Metalmecánica*\n\n¿Qué deseas hacer?",
        reply_markup={"inline_keyboard": [
            [{"text": "📊 Seguimiento a la Formación", "callback_data": "menu_seguimiento"}],
            [{"text": "📝 Plan Concertado", "callback_data": "menu_plan"}],
            [{"text": "📖 Consultar Reglamento", "callback_data": "menu_reglamento"}],
        ]}
    )

def menu_seguimiento(chat_id):
    set_estado(chat_id, {"modo": "seguimiento_menu"})
    send_message(chat_id,
        "📊 *Seguimiento a la Formación*\n\n¿Qué deseas gestionar?",
        reply_markup={"inline_keyboard": [
            [{"text": "📋 Llamado de Atención", "callback_data": "llamado_inicio"}],
            [{"text": "📈 Plan de Mejoramiento", "callback_data": "plan_mejoramiento"}],
            [{"text": "🏠 Menú principal", "callback_data": "menu_principal"}],
        ]}
    )

def menu_con_opciones(chat_id):
    """Muestra opciones al final de cada interacción"""
    send_message(chat_id,
        "¿Deseas hacer algo más?",
        reply_markup={"inline_keyboard": [
            [{"text": "📊 Nueva novedad", "callback_data": "menu_seguimiento"}],
            [{"text": "📝 Plan Concertado", "callback_data": "menu_plan"}],
            [{"text": "📖 Consultar Reglamento", "callback_data": "menu_reglamento"}],
            [{"text": "🏠 Menú principal", "callback_data": "menu_principal"}],
        ]}
    )

def consultar_groq(mensajes, system):
    try:
        messages = [{"role": "system", "content": system}] + mensajes
        r = requests.post(GROQ_URL,
            headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile", "messages": messages, "max_tokens": 800, "temperature": 0.2},
            timeout=30)
        data = r.json()
        if r.status_code != 200:
            logger.error(f"Groq error: {data}")
            return None
        return data["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"Groq exception: {e}")
        return None

# ─── FLUJO LLAMADO DE ATENCIÓN ────────────────────────────────────

def llamado_paso1(chat_id):
    set_estado(chat_id, {"modo": "llamado", "paso": "aprendiz"})
    send_message(chat_id, "📋 *Llamado de Atención*\n\n*Paso 1/6* — ¿Cuál es el *nombre completo* del aprendiz?")

def llamado_paso2(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "programa"
    set_estado(chat_id, e)
    send_message(chat_id,
        f"✅ Aprendiz: *{e['aprendiz']}*\n\n*Paso 2/6 — Programa de formación*",
        reply_markup={"inline_keyboard": [
            [{"text": "⚙️ CNC — Máquinas CNC", "callback_data": "prog_CNC"}],
            [{"text": "🔧 MEI — Electromecánico Industrial", "callback_data": "prog_MEI"}],
            [{"text": "🏭 MMI — Maquinaria Industrial", "callback_data": "prog_MMI"}],
            [{"text": "📊 GPI — Gestión Producción", "callback_data": "prog_GPI"}],
        ]})

def llamado_paso3(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "ficha"
    set_estado(chat_id, e)
    send_message(chat_id, f"✅ Programa: *{PROGRAMAS_NOMBRES.get(e['programa'],'?')}*\n\n*Paso 3/6* — Escribe el *número de ficha*:")

def llamado_paso3b(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "instructor"
    set_estado(chat_id, e)
    send_message(chat_id,
        f"✅ Ficha: *{e['ficha']}*\n\n*¿Qué instructor reporta la falta?*",
        reply_markup={"inline_keyboard": [
            [{"text": "Carlos Charris", "callback_data": "inst_Carlos Charris"}],
            [{"text": "Carlos Sabalza", "callback_data": "inst_Carlos Sabalza"}],
            [{"text": "Wilfrido Romero", "callback_data": "inst_Wilfrido Romero"}],
            [{"text": "Ana Maria Barrios", "callback_data": "inst_Ana Maria Barrios"}],
            [{"text": "Aramis Vitola", "callback_data": "inst_Aramis Vitola"}],
            [{"text": "Omar Pardey", "callback_data": "inst_Omar Pardey"}],
        ]})

def llamado_paso4(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "tipo"
    set_estado(chat_id, e)
    send_message(chat_id,
        f"✅ Instructor: *{e['instructor']}*\n\n*Paso 4/6 — Tipo de llamado*",
        reply_markup={"inline_keyboard": [
            [{"text": "🛡️ Disciplinario", "callback_data": "tipo_disciplinario"}],
            [{"text": "📚 Académico", "callback_data": "tipo_academico"}],
        ]})

def llamado_paso5(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "numero"
    set_estado(chat_id, e)
    send_message(chat_id,
        f"✅ Tipo: *{e['tipo_llamado'].capitalize()}*\n\n*Paso 5/6 — Número de llamado*",
        reply_markup={"inline_keyboard": [
            [{"text": "1️⃣ Primer Llamado", "callback_data": "num_primero"}],
            [{"text": "2️⃣ Segundo Llamado", "callback_data": "num_segundo"}],
        ]})

def llamado_paso6(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "falta"
    set_estado(chat_id, e)
    faltas = list(FALTAS_DISC.keys()) if e["tipo_llamado"] == "disciplinario" else list(FALTAS_ACAD.keys())
    botones = [[{"text": f, "callback_data": f"falta_{i}"}] for i, f in enumerate(faltas)]
    send_message(chat_id,
        f"✅ *{e['numero_llamado'].capitalize()} Llamado {e['tipo_llamado'].capitalize()}*\n\n*Paso 6/6 — Tipo de falta*",
        reply_markup={"inline_keyboard": botones})

def llamado_paso7(chat_id):
    e = get_estado(chat_id)
    falta = e.get("falta", "")
    if "inasistencia" in falta.lower() or "injustificada" in falta.lower():
        e["paso"] = "mes_falta"
        set_estado(chat_id, e)
        meses = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
        botones = [[{"text": m.capitalize(), "callback_data": f"mes_{m}"}] for m in meses]
        send_message(chat_id, f"✅ Falta: *{falta}*\n\n¿En qué mes?", reply_markup={"inline_keyboard": botones})
    else:
        e["paso"] = "detalles"
        set_estado(chat_id, e)
        send_message(chat_id, f"✅ Falta: *{falta}*\n\nDescribe brevemente la situación:")

def llamado_dias(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "dias_falta"
    set_estado(chat_id, e)
    send_message(chat_id, f"✅ Mes: *{e['mes_falta'].capitalize()}*\n\nEscribe los *días de inasistencia* separados por comas:\n_Ejemplo: 3, 10, 15, 22_")

def llamado_resumen(chat_id):
    e = get_estado(chat_id)
    n = e.get("numero_llamado","primer")
    t = e.get("tipo_llamado","disciplinario")
    resumen = (
        f"📋 *Resumen del Llamado de Atención*\n\n"
        f"• *Aprendiz:* {e.get('aprendiz','')}\n"
        f"• *Programa:* {PROGRAMAS_NOMBRES.get(e.get('programa',''),'')}\n"
        f"• *Ficha:* {e.get('ficha','')}\n"
        f"• *Instructor:* {e.get('instructor','')}\n"
        f"• *Tipo:* {t.capitalize()}\n"
        f"• *Número:* {n.capitalize()} llamado\n"
        f"• *Falta:* {e.get('falta','')}\n"
    )
    if e.get("mes_falta"):
        resumen += f"• *Mes:* {e['mes_falta'].capitalize()}\n"
        resumen += f"• *Días:* {e.get('dias_falta','')}\n"
    else:
        resumen += f"• *Descripción:* {e.get('descripcion','')}\n"
    resumen += "\n¿Generamos el documento?"
    send_message(chat_id, resumen,
        reply_markup={"inline_keyboard": [
            [{"text": "✅ Generar PDF", "callback_data": "generar_llamado"},
             {"text": "❌ Cancelar", "callback_data": "menu_principal"}],
        ]})

def llamado_generar_pdf(chat_id):
    e = get_estado(chat_id)
    send_typing(chat_id)
    send_message(chat_id, "⏳ Generando el Llamado de Atención...")
    try:
        e['programa_nombre'] = PROGRAMAS_NOMBRES.get(e.get('programa','CNC'), e.get('programa',''))
        pdf_bytes = generar_pdf_llamado_oficial(e)
        aprendiz = e.get("aprendiz", "Aprendiz")
        apellido = aprendiz.split()[-1]
        fecha = datetime.now().strftime("%Y%m%d")
        n = e.get("numero_llamado","primer")
        t = e.get("tipo_llamado","disciplinario")
        filename = f"Llamado_{n.capitalize()}_{t.capitalize()}_{apellido}_{fecha}.pdf"
        send_document(chat_id, pdf_bytes, filename,
            f"📄 {n.capitalize()} Llamado de Atención {t.capitalize()}\nAprendiz: {aprendiz}")
        send_message(chat_id,
            "✅ *Documento generado*\n\n"
            "Recuerda:\n"
            "• Imprimir y hacer firmar al aprendiz\n"
            "• Archivar en la carpeta del aprendiz\n"
            "• Notificar a coordinación académica")
        menu_con_opciones(chat_id)
    except Exception as ex:
        logger.error(f"Error PDF: {ex}")
        send_message(chat_id, f"❌ Error generando el documento. Intenta de nuevo.")
        menu_con_opciones(chat_id)

# ─── FLUJO PLAN CONCERTADO ────────────────────────────────────────

def plan_inicio(chat_id):
    set_estado(chat_id, {"modo": "plan", "paso": "archivo"})
    send_message(chat_id,
        "📝 *Plan Concertado*\n\n"
        "Por favor adjunta el archivo de *Juicios Evaluativos* (Excel .xlsx o .xls) de la ficha.\n\n"
        "El agente extraerá automáticamente:\n"
        "• Programa de formación\n"
        "• Número de ficha\n"
        "• Lista de aprendices EN FORMACION")

def plan_pedir_instructor(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "instructor"
    set_estado(chat_id, e)
    n = len(e.get("aprendices",[]))
    ficha = e.get("ficha_auto","")
    send_message(chat_id,
        f"✅ *{n} aprendices cargados*\n"
        f"📋 Ficha detectada: *{ficha}*\n\n"
        "¿Cuál es el *nombre del instructor*?")

def plan_pedir_fecha(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "fecha"
    set_estado(chat_id, e)
    send_message(chat_id,
        f"✅ Instructor: *{e.get('instructor','')}*\n\n"
        "¿Cuál es la *fecha del Plan Concertado*?\n"
        "_Escribe en formato DD/MM/YYYY — Ejemplo: 28/06/2026_")

def plan_detectar_programa(chat_id):
    """Detecta el programa desde el código extraído del Excel y va directo a proyectos"""
    e = get_estado(chat_id)
    codigo = e.get("codigo_programa", "")
    logger.info(f"Detectando programa para código: '{codigo}'")
    # Buscar programa en catálogo por código
    programa_encontrado = None
    for prog_key in CATALOGO:
        if codigo and codigo in prog_key:
            programa_encontrado = prog_key
            break
    logger.info(f"Programa encontrado: {programa_encontrado}")
    if programa_encontrado:
        e["programa_key"] = programa_encontrado
        set_estado(chat_id, e)
        plan_seleccionar_proyecto(chat_id)
    else:
        # No se detectó — mostrar lista para elegir manualmente
        send_message(chat_id, "No pude detectar el programa. Selecciona manualmente:")
        plan_seleccionar_programa(chat_id)

def plan_seleccionar_programa(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "programa"
    set_estado(chat_id, e)
    programas = list(CATALOGO.keys())
    botones = [[{"text": p[:60], "callback_data": f"plan_prog_{i}"}] for i, p in enumerate(programas)]
    send_message(chat_id,
        "📚 *Selecciona el Programa de Formación:*",
        reply_markup={"inline_keyboard": botones})

def plan_seleccionar_proyecto(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "proyecto"
    set_estado(chat_id, e)
    prog = e.get("programa_key","")
    proyectos = list(CATALOGO[prog]["proyectos"].keys())
    if len(proyectos) == 1:
        e["proyecto_key"] = proyectos[0]
        set_estado(chat_id, e)
        plan_seleccionar_fase(chat_id)
    else:
        botones = [[{"text": p[:60], "callback_data": f"plan_proy_{i}"}] for i, p in enumerate(proyectos)]
        send_message(chat_id,
            "📋 *Selecciona el Proyecto Formativo:*\n\n"
            "_(Hay más de un proyecto para este programa)_",
            reply_markup={"inline_keyboard": botones})

def plan_seleccionar_fase(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "fase"
    set_estado(chat_id, e)
    prog = e.get("programa_key","")
    proy = e.get("proyecto_key","")
    fases = list(CATALOGO[prog]["proyectos"][proy]["fases"].keys())
    botones = [[{"text": f[:60], "callback_data": f"plan_fase_{i}"}] for i, f in enumerate(fases)]
    send_message(chat_id,
        "🔄 *Selecciona la Fase del Proyecto:*",
        reply_markup={"inline_keyboard": botones})

def plan_seleccionar_actividad(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "actividad"
    set_estado(chat_id, e)
    prog = e.get("programa_key","")
    proy = e.get("proyecto_key","")
    fase = e.get("fase_key","")
    acts = list(CATALOGO[prog]["proyectos"][proy]["fases"][fase].keys())
    botones = [[{"text": a[:60], "callback_data": f"plan_act_{i}"}] for i, a in enumerate(acts)]
    send_message(chat_id,
        "🎯 *Selecciona la Actividad de Proyecto:*",
        reply_markup={"inline_keyboard": botones})

def plan_seleccionar_competencia(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "competencia"
    set_estado(chat_id, e)
    prog = e.get("programa_key","")
    proy = e.get("proyecto_key","")
    fase = e.get("fase_key","")
    act  = e.get("actividad_key","")
    comps = CATALOGO[prog]["proyectos"][proy]["fases"][fase][act]["competencias"]
    botones = [[{"text": c["nombre"][:60], "callback_data": f"plan_comp_{i}"}] for i, c in enumerate(comps)]
    send_message(chat_id,
        "🏆 *Selecciona la Competencia:*",
        reply_markup={"inline_keyboard": botones})

def plan_seleccionar_resultados(chat_id):
    e = get_estado(chat_id)
    e["paso"] = "resultados"
    if "resultados_sel" not in e:
        e["resultados_sel"] = []
    set_estado(chat_id, e)
    prog  = e.get("programa_key","")
    proy  = e.get("proyecto_key","")
    fase  = e.get("fase_key","")
    act   = e.get("actividad_key","")
    comp_i = e.get("competencia_idx",0)
    comps  = CATALOGO[prog]["proyectos"][proy]["fases"][fase][act]["competencias"]
    comp   = comps[comp_i]
    ras    = comp["resultados"]

    msg = "📚 *Resultados de Aprendizaje disponibles:*\n\n"
    for i, r in enumerate(ras):
        sel = "✅" if i in e["resultados_sel"] else "☐"
        msg += f"{sel} *{i+1}.* {r['ra'][:80]}...\n\n"

    botones = []
    for i, r in enumerate(ras):
        label = f"{'✅' if i in e['resultados_sel'] else '☐'} RA {i+1}"
        botones.append([{"text": label, "callback_data": f"plan_ra_{i}"}])

    botones.append([{"text": "✅ Confirmar selección y generar", "callback_data": "plan_generar"}])
    botones.append([{"text": "🏠 Cancelar", "callback_data": "menu_principal"}])

    send_message(chat_id, msg, reply_markup={"inline_keyboard": botones})

def plan_generar(chat_id):
    e = get_estado(chat_id)
    send_typing(chat_id)
    send_message(chat_id, "⏳ Generando los Planes Concertados...")
    try:
        prog  = e.get("programa_key","")
        proy  = e.get("proyecto_key","")
        fase  = e.get("fase_key","")
        act   = e.get("actividad_key","")
        comp_i = e.get("competencia_idx",0)
        comps  = CATALOGO[prog]["proyectos"][proy]["fases"][fase][act]["competencias"]
        comp   = comps[comp_i]
        ras    = comp["resultados"]
        sel_idx = e.get("resultados_sel",[])
        resultados_sel = [ras[i] for i in sel_idx] if sel_idx else ras

        aprendices = e.get("aprendices",[])
        proyecto_data = CATALOGO[prog]["proyectos"][proy]

        datos = {
            "programa":      prog,
            "instructor":    e.get("instructor","").upper(),
            "ficha":         e.get("ficha_auto",""),
            "proyecto":      proyecto_data["nombre_completo"],
            "fase":          fase,
            "observaciones": e.get("observaciones",""),
            "fecha_plan":    e.get("fecha_plan", datetime.now().strftime("%d/%m/%Y")),
        }

        if len(aprendices) == 1:
            pdf_bytes = generar_pdf_plan_concertado(aprendices, resultados_sel, datos)
            ap = aprendices[0]
            apellido = ap['nombre'].split()[-1]
            filename = f"PlanConcertado_{apellido}_{datos['ficha']}.pdf"
            send_document(chat_id, pdf_bytes, filename,
                f"📝 Plan Concertado\nAprendiz: {ap['nombre']}\nFicha: {datos['ficha']}")
        else:
            zip_bytes = generar_zip_plan_concertado(aprendices, resultados_sel, datos)
            pdf_bytes = generar_pdf_plan_concertado(aprendices, resultados_sel, datos)
            filename_zip = f"PlanConcertado_Ficha_{datos['ficha']}.zip"
            filename_pdf = f"PlanConcertado_Ficha_{datos['ficha']}.pdf"
            send_document(chat_id, zip_bytes, filename_zip,
                f"📦 ZIP con {len(aprendices)} Planes Concertados individuales — Ficha {datos['ficha']}")
            send_document(chat_id, pdf_bytes, filename_pdf,
                f"📄 PDF único con {len(aprendices)} páginas — Ficha {datos['ficha']}")

        send_message(chat_id, f"✅ *{len(aprendices)} Plan(es) Concertado(s) generado(s)*")
        menu_con_opciones(chat_id)
    except Exception as ex:
        logger.error(f"Error Plan Concertado: {ex}")
        send_message(chat_id, f"❌ Error generando. Intenta de nuevo.\n`{str(ex)[:100]}`")
        menu_con_opciones(chat_id)

def procesar_excel_aprendices(file_bytes, filename="file.xlsx"):
    """Lee aprendices del Excel de Juicios Evaluativos.
    Flexible: detecta encabezado en cualquier fila, soporta .xls y .xlsx,
    acepta cualquier variación de tildes/mayúsculas en el estado.
    """
    import io, unicodedata as _ud

    def norm(s):
        """Normaliza texto: sin tildes, sin espacios extra, mayúsculas"""
        s = str(s or '').strip()
        return ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn').upper()

    try:
        buf = io.BytesIO(file_bytes)
        fname_lower = filename.lower()

        # ── Leer todas las filas según formato ────────────────────
        if fname_lower.endswith('.xls') and not fname_lower.endswith('.xlsx'):
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            rows = []
            for rx in range(ws_xls.nrows):
                row = []
                for cx in range(ws_xls.ncols):
                    cell = ws_xls.cell(rx, cx)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        v = cell.value
                        row.append(int(v) if v == int(v) else v)
                    else:
                        row.append(cell.value)
                rows.append(tuple(row))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

        # ── Detectar ficha y código de programa en primeras filas ──
        ficha_detectada = None
        codigo_programa_detectado = None
        for row in rows[:20]:
            vals_norm = [norm(c) for c in row]
            for i, v in enumerate(vals_norm):
                if 'FICHA' in v and not ficha_detectada:
                    for c in row:
                        s = str(c or '').replace('.0','').strip()
                        if s.isdigit() and len(s) >= 5:
                            ficha_detectada = s
                            break
                # Detectar código del programa: busca celda con "código" o "cógigo" o variantes
                if ('COGIGO' in v or 'CODIGO' in v or 'COGIDO' in v or 'GIGO' in v) and not codigo_programa_detectado:
                    for c in row:
                        s = str(c or '').replace('.0','').strip()
                        # Código de programa: 5-7 dígitos
                        if s.isdigit() and 5 <= len(s) <= 7:
                            codigo_programa_detectado = s
                            break

        # ── Detectar fila de encabezado de forma flexible ─────────
        # Busca la fila que tenga columnas de datos de aprendices
        # Indicadores: "NOMBRE", "ESTADO", "DOCUMENTO", "TIPO"
        header_row = None
        best_score = 0
        for i, row in enumerate(rows):
            vals_norm = [norm(c) for c in row]
            score = 0
            if any('NOMBRE' in v for v in vals_norm): score += 3
            if any('ESTADO' in v for v in vals_norm): score += 3
            if any('DOCUMENTO' in v for v in vals_norm): score += 2
            if any('APELLIDO' in v for v in vals_norm): score += 2
            if any('TIPO' in v for v in vals_norm): score += 1
            if score > best_score:
                best_score = score
                header_row = i

        if header_row is None or best_score < 3:
            logger.error(f"No se detectó encabezado (score={best_score})")
            return [], ficha_detectada

        headers_norm = [norm(c) for c in rows[header_row]]
        logger.info(f"Encabezado detectado en fila {header_row}: {headers_norm[:8]}")

        # ── Mapear columnas de forma flexible ─────────────────────
        col_tipo = col_num = col_nombre = col_apellido = col_estado = None
        for i, h in enumerate(headers_norm):
            if col_tipo    is None and 'TIPO' in h and 'DOCUMENTO' in h: col_tipo = i
            if col_num     is None and ('NUMERO' in h or 'DOCUMENTO' in h) and 'TIPO' not in h: col_num = i
            if col_nombre  is None and h == 'NOMBRE': col_nombre = i
            if col_apellido is None and 'APELLIDO' in h: col_apellido = i
            if col_estado  is None and 'ESTADO' in h: col_estado = i

        # Fallback: buscar "nombre" parcial si no se encontró exacto
        if col_nombre is None:
            for i, h in enumerate(headers_norm):
                if 'NOMBRE' in h:
                    col_nombre = i
                    break

        if col_nombre is None:
            logger.error("No se encontró columna Nombre")
            return [], ficha_detectada

        # ── Extraer aprendices únicos EN FORMACION ─────────────────
        vistos = set()
        aprendices = []

        for row in rows[header_row + 1:]:
            try:
                nombre = str(row[col_nombre] or '').strip() if col_nombre is not None and len(row) > col_nombre else ''
                if not nombre or norm(nombre) in ('', 'NOMBRE', 'NAN'):
                    continue

                # Filtrar por estado si existe columna
                if col_estado is not None and len(row) > col_estado:
                    estado_norm = norm(row[col_estado])
                    # Acepta cualquier variación: EN FORMACION, En Formación, en formacion, etc.
                    if 'FORMAC' not in estado_norm:
                        continue

                apellido = str(row[col_apellido] or '').strip() if col_apellido is not None and len(row) > col_apellido else ''
                nombre_completo = f"{nombre} {apellido}".strip()

                tipo_doc = str(row[col_tipo] or 'CC').strip() if col_tipo is not None and len(row) > col_tipo else 'CC'
                num_doc  = str(row[col_num]  or '').strip().replace('.0','') if col_num is not None and len(row) > col_num else ''

                # Deduplicar por número de documento
                clave = num_doc if num_doc else nombre_completo
                if clave in vistos:
                    continue
                vistos.add(clave)

                aprendices.append({'nombre': nombre_completo, 'doc': f"{tipo_doc} {num_doc}".strip()})

            except Exception as ex:
                logger.warning(f"Error en fila: {ex}")
                continue

        logger.info(f"Aprendices encontrados: {len(aprendices)}, ficha: {ficha_detectada}")
        return aprendices, ficha_detectada, codigo_programa_detectado

    except Exception as ex:
        logger.error(f"Error leyendo Excel: {ex}")
        return [], None, None


@app.route("/webhook", methods=["POST"])
def webhook():
    try:
        data = request.get_json()

        # ── CALLBACKS (botones) ───────────────────────────────────
        if "callback_query" in data:
            cb = data["callback_query"]
            chat_id = cb["message"]["chat"]["id"]
            cb_data = cb["data"]
            answer_callback(cb["id"])
            e = get_estado(chat_id)

            if cb_data == "menu_principal":
                menu_principal(chat_id)
            elif cb_data == "menu_seguimiento":
                menu_seguimiento(chat_id)
            elif cb_data == "llamado_inicio":
                llamado_paso1(chat_id)
            elif cb_data == "plan_mejoramiento":
                set_estado(chat_id, {"modo": "seguimiento", "historial": []})
                send_message(chat_id,
                    "📈 *Plan de Mejoramiento*\n\n"
                    "Descríbeme la situación del aprendiz y te indico el procedimiento:\n\n"
                    "_Ejemplo: Aprendiz Carlos Gómez, ficha PCMCNC-5, nota de 2.5 en soldadura_")
            elif cb_data == "menu_reglamento":
                set_estado(chat_id, {"modo": "reglamento", "historial": []})
                send_message(chat_id,
                    "📖 *Consulta del Reglamento*\n\n"
                    "¿Qué deseas consultar del Acuerdo 0009 de 2024?")
            elif cb_data == "menu_plan":
                plan_inicio(chat_id)
            elif cb_data == "nueva_novedad":
                menu_seguimiento(chat_id)

            # Llamado — programa
            elif cb_data.startswith("prog_"):
                prog = cb_data.replace("prog_","")
                e["programa"] = prog
                set_estado(chat_id, e)
                llamado_paso3(chat_id)
            # Llamado — instructor
            elif cb_data.startswith("inst_"):
                inst = cb_data.replace("inst_","")
                e["instructor"] = inst
                set_estado(chat_id, e)
                llamado_paso4(chat_id)
            # Llamado — tipo
            elif cb_data.startswith("tipo_"):
                e["tipo_llamado"] = cb_data.replace("tipo_","")
                set_estado(chat_id, e)
                llamado_paso5(chat_id)
            # Llamado — número
            elif cb_data.startswith("num_"):
                e["numero_llamado"] = cb_data.replace("num_","")
                set_estado(chat_id, e)
                llamado_paso6(chat_id)
            # Llamado — falta
            elif cb_data.startswith("falta_"):
                idx = int(cb_data.replace("falta_",""))
                pool = list(FALTAS_DISC.keys()) if e.get("tipo_llamado") == "disciplinario" else list(FALTAS_ACAD.keys())
                e["falta"] = pool[idx]
                set_estado(chat_id, e)
                llamado_paso7(chat_id)
            # Llamado — mes
            elif cb_data.startswith("mes_"):
                e["mes_falta"] = cb_data.replace("mes_","")
                set_estado(chat_id, e)
                llamado_dias(chat_id)
            # Llamado — generar
            elif cb_data == "generar_llamado":
                llamado_generar_pdf(chat_id)

            # Plan Concertado — programa
            elif cb_data.startswith("plan_prog_"):
                idx = int(cb_data.replace("plan_prog_",""))
                prog_key = list(CATALOGO.keys())[idx]
                e["programa_key"] = prog_key
                set_estado(chat_id, e)
                plan_seleccionar_proyecto(chat_id)
            # Plan Concertado — proyecto
            elif cb_data.startswith("plan_proy_"):
                idx = int(cb_data.replace("plan_proy_",""))
                prog_key = e.get("programa_key","")
                proy_key = list(CATALOGO[prog_key]["proyectos"].keys())[idx]
                e["proyecto_key"] = proy_key
                set_estado(chat_id, e)
                plan_seleccionar_fase(chat_id)
            # Plan Concertado — fase
            elif cb_data.startswith("plan_fase_"):
                idx = int(cb_data.replace("plan_fase_",""))
                prog = e.get("programa_key","")
                proy = e.get("proyecto_key","")
                fase_key = list(CATALOGO[prog]["proyectos"][proy]["fases"].keys())[idx]
                e["fase_key"] = fase_key
                set_estado(chat_id, e)
                plan_seleccionar_actividad(chat_id)
            # Plan Concertado — actividad
            elif cb_data.startswith("plan_act_"):
                idx = int(cb_data.replace("plan_act_",""))
                prog = e.get("programa_key","")
                proy = e.get("proyecto_key","")
                fase = e.get("fase_key","")
                act_key = list(CATALOGO[prog]["proyectos"][proy]["fases"][fase].keys())[idx]
                e["actividad_key"] = act_key
                set_estado(chat_id, e)
                plan_seleccionar_competencia(chat_id)
            # Plan Concertado — competencia
            elif cb_data.startswith("plan_comp_"):
                idx = int(cb_data.replace("plan_comp_",""))
                e["competencia_idx"] = idx
                e["resultados_sel"] = []
                set_estado(chat_id, e)
                plan_seleccionar_resultados(chat_id)
            # Plan Concertado — resultado de aprendizaje (toggle)
            elif cb_data.startswith("plan_ra_"):
                idx = int(cb_data.replace("plan_ra_",""))
                sel = e.get("resultados_sel", [])
                if idx in sel:
                    sel.remove(idx)
                else:
                    sel.append(idx)
                e["resultados_sel"] = sel
                set_estado(chat_id, e)
                plan_seleccionar_resultados(chat_id)
            # Plan Concertado — generar
            elif cb_data == "plan_generar":
                plan_generar(chat_id)

            return jsonify({"ok": True})

        # ── MENSAJES DE TEXTO ─────────────────────────────────────
        if "message" not in data:
            return jsonify({"ok": True})

        message = data["message"]
        chat_id = message["chat"]["id"]
        text    = message.get("text", "")
        username = message.get("from",{}).get("username","")

        # Documentos (Excel para Plan Concertado)
        if "document" in message:
            doc = message["document"]
            fname = doc.get("file_name","")
            e = get_estado(chat_id)
            if e.get("modo") == "plan" and e.get("paso") == "archivo":
                if fname.lower().endswith(('.xlsx','.xls')):
                    send_typing(chat_id)
                    send_message(chat_id, "⏳ Leyendo el archivo de Juicios Evaluativos...")
                    try:
                        file_id = doc["file_id"]
                        r = requests.get(f"{TELEGRAM_API}/getFile?file_id={file_id}", timeout=10)
                        file_path = r.json()["result"]["file_path"]
                        r2 = requests.get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}", timeout=30)
                        file_bytes = r2.content
                        aprendices, ficha, codigo = procesar_excel_aprendices(file_bytes, fname)
                        if aprendices:
                            e["aprendices"] = aprendices
                            e["ficha_auto"] = ficha or "Sin detectar"
                            set_estado(chat_id, e)
                            plan_pedir_instructor(chat_id)
                        else:
                            send_message(chat_id, "❌ No se encontraron aprendices EN FORMACION en el archivo. Verifica el formato e intenta de nuevo.")
                    except Exception as ex:
                        logger.error(f"Error archivo: {ex}")
                        send_message(chat_id, f"❌ Error leyendo el archivo: `{str(ex)[:100]}`\n\nIntenta de nuevo.")
                else:
                    send_message(chat_id, "❌ Por favor adjunta un archivo Excel (.xlsx o .xls)")
            return jsonify({"ok": True})

        if not text:
            return jsonify({"ok": True})

        logger.info(f"Msg @{username}: {text[:60]}")

        if text in ["/start", "/menu"]:
            menu_principal(chat_id)
            return jsonify({"ok": True})
        if text == "/nuevo":
            menu_principal(chat_id)
            return jsonify({"ok": True})

        e = get_estado(chat_id)
        modo = e.get("modo","menu")
        paso = e.get("paso","")

        # ── FLUJO LLAMADO ──
        if modo == "llamado":
            if paso == "aprendiz":
                e["aprendiz"] = text.strip().title()
                set_estado(chat_id, e)
                llamado_paso2(chat_id)
            elif paso == "ficha":
                e["ficha"] = text.strip()
                set_estado(chat_id, e)
                llamado_paso3b(chat_id)
            elif paso == "detalles":
                e["descripcion"] = text.strip()
                set_estado(chat_id, e)
                llamado_resumen(chat_id)
            elif paso == "dias_falta":
                e["dias_falta"] = text.strip()
                set_estado(chat_id, e)
                llamado_resumen(chat_id)

        # ── FLUJO PLAN CONCERTADO ──
        elif modo == "plan":
            if paso == "archivo":
                send_message(chat_id, "📎 Por favor adjunta el archivo Excel (.xlsx) de Juicios Evaluativos.")
            elif paso == "instructor":
                e["instructor"] = text.strip()
                set_estado(chat_id, e)
                plan_pedir_fecha(chat_id)
            elif paso == "fecha":
                e["fecha_plan"] = text.strip()
                set_estado(chat_id, e)
                plan_detectar_programa(chat_id)

        # ── FLUJO SEGUIMIENTO (análisis IA) ──
        elif modo in ("seguimiento","seguimiento_menu"):
            if "historial" not in e:
                e["historial"] = []
            e["historial"].append({"role":"user","content":text})
            set_estado(chat_id, e)
            send_typing(chat_id)
            respuesta = consultar_groq(e["historial"], SYSTEM_ANALISIS)
            if respuesta:
                e["historial"].append({"role":"assistant","content":respuesta})
                set_estado(chat_id, e)
                send_message(chat_id, respuesta,
                    reply_markup={"inline_keyboard":[
                        [{"text":"📋 Generar Llamado de Atención","callback_data":"llamado_inicio"}],
                        [{"text":"🔄 Nueva novedad","callback_data":"nueva_novedad"},
                         {"text":"🏠 Menú","callback_data":"menu_principal"}],
                    ]})
            else:
                send_message(chat_id,"Error procesando. Intenta de nuevo.")
                menu_con_opciones(chat_id)

        # ── FLUJO REGLAMENTO ──
        elif modo == "reglamento":
            if "historial" not in e:
                e["historial"] = []
            e["historial"].append({"role":"user","content":text})
            set_estado(chat_id, e)
            send_typing(chat_id)
            respuesta = consultar_groq(e["historial"], SYSTEM_REGLAMENTO)
            if respuesta:
                e["historial"].append({"role":"assistant","content":respuesta})
                set_estado(chat_id, e)
                send_message(chat_id, respuesta,
                    reply_markup={"inline_keyboard":[
                        [{"text":"❓ Otra consulta","callback_data":"menu_reglamento"}],
                        [{"text":"🏠 Menú principal","callback_data":"menu_principal"}],
                    ]})
            else:
                send_message(chat_id,"Error. Intenta de nuevo.")
                menu_con_opciones(chat_id)

        # ── SIN FLUJO — interpretar intención ──
        else:
            txt = text.lower()
            if "llamado" in txt:
                llamado_paso1(chat_id)
            elif "plan concertado" in txt or "concertado" in txt:
                plan_inicio(chat_id)
            elif "reglamento" in txt or "acuerdo" in txt:
                set_estado(chat_id, {"modo":"reglamento","historial":[{"role":"user","content":text}]})
                send_typing(chat_id)
                r = consultar_groq([{"role":"user","content":text}], SYSTEM_REGLAMENTO)
                if r:
                    send_message(chat_id, r, reply_markup={"inline_keyboard":[[{"text":"🏠 Menú","callback_data":"menu_principal"}]]})
                    menu_con_opciones(chat_id)
            else:
                set_estado(chat_id, {"modo":"seguimiento","historial":[{"role":"user","content":text}]})
                send_typing(chat_id)
                r = consultar_groq([{"role":"user","content":text}], SYSTEM_ANALISIS)
                if r:
                    send_message(chat_id, r, reply_markup={"inline_keyboard":[
                        [{"text":"📋 Generar Llamado","callback_data":"llamado_inicio"}],
                        [{"text":"🏠 Menú","callback_data":"menu_principal"}],
                    ]})
                    menu_con_opciones(chat_id)

        return jsonify({"ok": True})

    except Exception as ex:
        logger.error(f"Error webhook: {ex}")
        return jsonify({"ok": True})

@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "Agente SENA activo", "bot": "@AgenteMecanizadoSena_bot"})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
